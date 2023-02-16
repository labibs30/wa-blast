import openpyxl as excel
import urllib
import time
import random
import string
from pyparsing import col
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.chrome.options import Options

import PySimpleGUI as sg
import openpyxl
import pandas as pd
import os

# BLAST AKA BECK END-Nya
PATH = "./chromedriver.exe"
driver = webdriver.Chrome()

def element_presence(by, xpath, time):
    element_present = EC.presence_of_element_located((By.XPATH, xpath))
    WebDriverWait(driver, time).until(element_present)

def sendMessage(contact, message, sendDelay):
    XPATH = '//*[@id="main"]/footer/div[1]/div/span[2]/div/div[2]/div[1]/div/div[1]/p'

    print("Sending message for {0}".format(contact))
    url = "https://web.whatsapp.com/send?phone={0}&text={1}".format(contact, urllib.parse.quote(message))
    driver.get(url)

    print("Finding Input Box...")
    element_presence(By.XPATH, XPATH , 30)
    inputBox = driver.find_element(By.XPATH, XPATH)
    
    if inputBox != None: print("Input Box found!")
    else: print("Input Box not found!")
    
    inputBox.send_keys(Keys.ENTER)
    print("Message sent.")
    print("======================================================================")

    time.sleep(sendDelay)

def readContact(fileName):
    #Load File Excel
    file = excel.load_workbook(fileName, data_only=True)
    
    #Load Data Setting
    settings = file['Setting']
    contactColumn = settings['B2'].value
    messageColumn = settings['B3'].value
    sendDelay = settings['B4'].value
    sendTotal = settings['B5'].value
    
    #Load Database
    database = file['Database']
    contactData = database[contactColumn]
    messageData = database[messageColumn]
    
    for cell in range(1, sendTotal+1):
        contact = str(contactData[cell].value)
        message = str(messageData[cell].value)
        print(message)
        sendMessage(contact, message, sendDelay)
    
    # time.sleep(30)
def readContact(fileName):
    #Load File Excel
    file = excel.load_workbook(fileName, data_only=True)
    
    #Load Data Setting
    settings = file['Setting']
    contactColumn = settings['B2'].value
    messageColumn = settings['B3'].value
    sendDelay = settings['B4'].value
    sendInit = settings['B5'].value
    sendFinal = settings['B6'].value
    
    #Load Database
    database = file['Database']
    contactData = database[contactColumn]
    messageData = database[messageColumn]
    
    for cell in range(sendInit, sendFinal):
        contact = str(contactData[cell].value)
        message = str(messageData[cell].value)
        # message = "Halo bos apa kabarnya"
        sendMessage(contact, message, sendDelay)
    
    # time.sleep(30)


# GUI AKA FRONT END-Nya
sg.theme('DarkTeal9')   # Add a touch of color

if os.path.exists('database.xlsx'):
    pass
else:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    database = workbook.create_sheet("Database")
    setting = workbook.create_sheet("Setting")
    database['A1'] = 'no'
    database['B1'] = 'name'
    database['C1'] = 'kontak'
    database['D1'] = 'message'
    setting['A1'] = 'variable'
    setting['A2'] = 'Kolom Nomor'
    setting['A3'] = 'Kolom Pesan'
    setting['A4'] = 'Jeda Kirim'
    setting['A5'] = 'init'
    setting['A6'] = 'final'
    setting['B1'] = 'value'
    setting['B2'] = 'C'
    setting['B3'] = 'D'
    setting['B4'] = 4
    setting['B5'] = 0
    setting['B6'] = 0

    
    workbook.save("database.xlsx")
    
size = (400,300)
# All the stuff inside your window.
tab1_layout = [  [sg.Text('Name:', size=(15, 1)), sg.InputText()],
            [sg.Text('Contact', size=(15, 1)), sg.InputText()],
            [sg.Submit(), sg.Cancel() ] 
        ]

tab2_layout = [
    [sg.Text('''
             Apabila ingin menggunakan nama pengguna 
             masukkan syntax "&B2&"
            ''')],
    [sg.Text('Pesan:', size=(15, 1)),sg.Multiline(size=(30, 5), key='message'), sg.InputText(size=(40, 10)), ],
    [sg.Button('Post'), sg.Button('Clear')] 
]

tab3_layout = [
    [sg.Text('Baris Awal:', size=(15, 1), ),  sg.InputText(key='init')],
    [sg.Text('Bais Akhir:', size=(15, 1), ), sg.InputText(key='final')],
    [sg.Button('Set'), sg.Button('Clear')] 
]

tab4_layout = [
    [sg.Button('Send',  size=(15, 1))] 
]
layout = [  
          [sg.TabGroup([
              [sg.Tab('Input Contact', tab1_layout, key='in_c'), 
               sg.Tab('Input Message', tab2_layout, key='in_m'), 
               sg.Tab('Setting', tab3_layout, 'set'), 
               sg.Tab('Send', tab4_layout, 'send')]
             ])
          ]
        ]
# Create the Window
window = sg.Window('Window Title', layout, size=size)
# Event Loop to process "events" and get the "values" of the inputs
def clear_input():
    for key in values:
        window[key]('')
    return None
           
while True:
    event, values = window.read()
    workbook = openpyxl.load_workbook("database.xlsx")
    db = workbook["Database"]
    st = workbook["Setting"]
    next_row_db = db.max_row + 1 
    next_row_st = st.max_row + 1 
    total = next_row_db
    if event == sg.WIN_CLOSED:   # if user closes window or clicks cancel
        break
    
    if event == 'Submit':
       db.cell(row=next_row_db, column=1, value=next_row_db-1)
       db.cell(row=next_row_db, column=2, value=values[0])
       db.cell(row=next_row_db, column=3, value=int(values[1]))
       workbook.save("database.xlsx")
       sg.popup("Data saved successfully!")
       clear_input()
    if event == 'Post':
        for item in range(2, next_row_db):
            db.cell(row=item, column=4, value='="'+values['message']+'"')
        # st.cell(row=4, column=2, value=values['message'])
        workbook.save("database.xlsx")
        clear_input()       
        
        sg.popup("Data saved successfully!") 
    
    if event == 'Set':
        st.cell(row=5, column=2, value=int(values['init']))
        st.cell(row=6, column=2, value=int(values['final']))
        workbook.save("database.xlsx")
        sg.popup("Data saved successfully!") 
        clear_input()       
        
        
    if event == 'Clear':
        clear_input()
    
    if event == 'Send':
        target = readContact("./database.xlsx")
    
window.close()
